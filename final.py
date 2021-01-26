from pandas import ExcelFile
from openpyxl import load_workbook
from selenium import webdriver
from bs4 import BeautifulSoup
from shutil import move
from re import split
from glob import glob

from datetime import datetime
now = datetime.now()
currentTime = now.strftime("%H.%M")

#TODO: Load SMCC data sheet
#NOTE: Lấy tất cả xlsx file trong folder demo
sources = glob("./source/*.xlsx")
files = ["", "", ""]
for file in sources:
    if "reputa" in file.lower():
        files[2] = file
    elif "smcctemplate" in file.lower():
        files[1] = file
    else: files[0] = file
#NOTE: SMCC sheet
wb1 = load_workbook(files[0])
sh1 = wb1.active
#NOTE: SMCC template sheet when no SMCC sheet
wb2 = load_workbook(files[1])
sh2 = wb2.active

max = sh1.max_row

#TODO: turn Reputa sheet into dataframe and modify to fit SMCC sheet
excels = ExcelFile(files[2])
frames = excels.parse(excels.sheet_names[0], header=6,index_col=None)
newframes = frames[["STT", "Ngày", "Thời gian", "Tiêu đề", "URL", "Tóm Tắt", "Sắc thái", "Like", "Comment", "Share", "Tên miền"]].copy() #đổi thứ tự cột
newframes["STT"] = newframes["STT"].apply(lambda x: x + max - 2) #sửa STT
newframes["Thời gian"] = newframes["Thời gian"].apply(lambda x: ":".join(str(x).split(":")[:2])) #sửa thời gian
newframes["Sắc thái"] = newframes["Sắc thái"].apply(lambda x: str(x).capitalize()) #sửa sắc thái
newframes["Sắc thái"] = newframes["Sắc thái"].replace("Trung lập", "Trung tính")

#NOTE: sửa tác giả
def getAuthorUrl(link):
    a = split("://", link)
    newlink = a[0] + "://m." + a[1]
    op = webdriver.ChromeOptions()
    op.add_argument('headless')
    driver = webdriver.Chrome(options=op)
    driver.get(newlink)
    page = driver.page_source
    soup = BeautifulSoup(page, "html.parser")
    if soup.find_all("div", class_="_5rgr async_like") == []:
        author_url = "Co vde"
    else: 
        names = soup.find_all("div", class_="_5rgr async_like")[0]
        driver.close()
        # names = soup.find_all('a')['href']
        id = split(",|:", str(names))
        author_url = "http://facebook.com/{}".format(id[3][1:-1])
        return author_url

#thêm cột tác giả
list_tacgia = []
for count, value in enumerate(newframes["Tên miền"]):
        # print(value)
    if value == "facebook.com":
        author_url = getAuthorUrl(newframes["URL"][count])
        print(author_url)
        list_tacgia.append(author_url)
    else:
        list_tacgia.append("http://{}".format(value))
newframes["Tác giả"] = list_tacgia 

#thêm cột phân loại
def getPostUrl(link):
    op = webdriver.ChromeOptions()
    op.add_argument('headless')
    driver = webdriver.Chrome(options=op)
    driver.get(link)
    url = driver.current_url
    driver.close()
    print(url)
    return url

list_phanloai = []
for count, value in enumerate(frames["Nguồn"]):
    if value == "Khác":
        list_phanloai.append("Báo")
    elif value == "Forum":
        list_phanloai.append("Diễn đàn")
    elif value == "Mạng xã hội":
        url = getPostUrl(newframes["URL"][count])
        if "groups" in url:
            list_phanloai.append("Bài đăng nhóm")
        else:
            list_phanloai.append("Bài đăng fanpage")
    else:
        continue
newframes["Phân loại"] = list_phanloai

#TODO: Chuyển dataframe thành list và append vào SMCC sheet
data = newframes.values.tolist()
if max != 2:
    for i, j in enumerate(data):
        sh1.append(j)
        sh1.row_dimensions[max+i].height = 21
    wb1.save("result/báo_cáo_Huawei Media Monitoring_{}_{}.xlsx".format(newframes["Ngày"][0], currentTime))
else:
    for i, j in enumerate(data):
        sh2.append(j)
        sh2.row_dimensions[max+i+1].height = 21
    wb2.save("result/báo_cáo_Huawei Media Monitoring_{}_{}.xlsx".format(newframes["Ngày"][0], currentTime))

#TODO: Sắp xếp lại file
#NOTE: Move file đã match vào folder data
move(files[0], "./data")
move(files[2], "./data")

